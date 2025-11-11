# pl_excel.py
# Генерация ПЛ по шаблону shablon.xlsx
# Требует: pip install openpyxl

from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook

PL_TEMPLATE_PATH = Path("excel/shablon.xlsx")
def get_default_output_dir():
    # По умолчанию: папка программы
    return Path.cwd() / "Путевые листы"


def fmt_dt(iso_or_str):
    """Формат: ДД.ММ.ГГГГ ЧЧ:ММ. Поддерживает ISO-строки, обрезает TZ."""
    if not iso_or_str:
        return ""
    try:
        s = str(iso_or_str)
        s = s.split('Z')[0]
        if '+' in s:
            s = s.split('+')[0]
        dt = datetime.fromisoformat(s)
        return dt.strftime("%d.%m.%Y")
    except Exception:
        return str(iso_or_str)


def fmt_num(val):
    if val is None:
        return ""
    return str(val)


def render_driver_or_mechanic(driver2_obj):
    """
    Если у второго водителя поле 'driver_license' похоже на номер (есть хотя бы одна цифра),
    возвращает 'Водитель 2'. Иначе — 'Механик'. Слова вроде 'механик', 'нет', '—' тоже трактуются как механик.
    """
    if not isinstance(driver2_obj, dict):
        return "Механик"
    raw = (driver2_obj.get("driver_license") or "").strip()

    if not raw:
        return "Механик"

    # Нормализуем для проверок
    raw_lower = raw.lower()

    # Частые словесные значения, явно не номер
    non_number_words = {"механик", "нет", "без", "отсутствует", "н/д", "—", "-", "no", "none"}
    if raw_lower in non_number_words:
        return "Механик"

    # Валидный номер должен содержать хотя бы одну цифру
    has_digit = any(ch.isdigit() for ch in raw)
    if not has_digit:
        return "Механик"

    # Разрешим символы номера: буквы/цифры/пробел/дефис/№ (простая эвристика)
    allowed = set("abcdefghijklmnopqrstuvwxyzабвгдеёжзийклмнопрстуфхцчшщъыьэюя0123456789 -№")
    if all((c.lower() in allowed) for c in raw):
        return "Водитель 2"

    # Если есть странные символы, но присутствуют цифры — всё равно считаем номером
    return "Водитель 2"



def build_context(payload, dictionaries):
    """
    payload — данные созданного ПЛ (id-ссылки, строки, даты без TZ).
    dictionaries — словари по id: drivers, cars, podryads, gruzes, loading-points,
                   unloading-points, organizations, customers, seasons,
                   car-markas, car-models, default_pl_settings (кастомный).
    """
    # Из payload
    numberPL = payload.get("numberPL", "")
    marsh = payload.get("marsh", "")
    dispatch_info = payload.get("dispatch_info", "")
    comment = payload.get("comment", "")
    numberTN = payload.get("numberTN", "")
    tonn = payload.get("tonn", "")
    fuel_consumption = payload.get("fuel_consumption", "")

    # Даты (payload обычно без TZ); {dataPOPL} по вашему ТЗ — текущее время генерации
    dataPOPL = fmt_dt(payload.get("dataPOPL"))
    dataSDPL = fmt_dt(payload.get("dataSDPL"))
    loading_time = fmt_dt(payload.get("loading_time"))
    unloading_time = fmt_dt(payload.get("unloading_time"))

    # +10 дней
    data_to = (datetime.now() + timedelta(days=10)).strftime("%d.%m.%Y")

    # Справочники
    drivers = dictionaries.get("drivers", {})
    cars = dictionaries.get("cars", {})
    podryads = dictionaries.get("podryads", {})
    gruzes = dictionaries.get("gruzes", {})
    loading_points = dictionaries.get("loading-points", {})
    unloading_points = dictionaries.get("unloading-points", {})
    organizations = dictionaries.get("organizations", {})
    customers = dictionaries.get("customers", {})
    seasons = dictionaries.get("seasons", {})
    car_markas = dictionaries.get("car-markas", {})
    car_models = dictionaries.get("car-models", {})

    defaults = dictionaries.get("default_pl_settings", {}) or {}
    # Дистанция и диспетчер из настроек
    distance_val = defaults.get("distance", "") if payload.get("distance") in [None, "", []] else payload.get("distance")
    dispatcher_val = defaults.get("dispatcher", "")  # ожидаем, что вы добавите 'dispatcher' в настройки

    # По id
    driver_id = payload.get("driver")
    driver = drivers.get(driver_id, {})
    driver2_id = payload.get("driver2")
    driver2 = drivers.get(driver2_id, {})
    car_id = payload.get("number")
    car = cars.get(car_id, {})
    pod_id = payload.get("pod")
    pod = podryads.get(pod_id, {})
    gruz_id = payload.get("gruz")
    gruz = gruzes.get(gruz_id, {})

    lp_id = payload.get("loading_point")
    up_id = payload.get("unloading_point")
    org_id = payload.get("organization")
    cust_id = payload.get("customer")
    season_id = payload.get("season")

    # Марка по id из car['marka']
    marka_name = ""
    car_marka_id = car.get("marka")
    if car_marka_id in car_markas:
        marka_name = car_markas[car_marka_id].get("name", "")

    context = {
        "{numberPL}": numberPL,
        "{marsh}": marsh,

        "{driver_full_name}": driver.get("full_name", ""),
        "{driver2_full_name}": driver2.get("full_name", ""),
        "{driver_snils}": driver.get("snils", ""),
        "{driver2_snils}": driver2.get("snils", ""),
        "{driver_license}": driver.get("driver_license", ""),
        "{driver2_license}": driver2.get("driver_license", ""),

        "{car_number}": car.get("number", ""),
        "{car_number_pr}": car.get("number_pr", ""),
        "{brand}": marka_name,

        "{contractor_name}": pod.get("org_name", ""),
        "{gruz_name}": gruz.get("name", ""),

        "{loading_point}": loading_points.get(lp_id, {}).get("name", ""),
        "{unloading_point}": unloading_points.get(up_id, {}).get("name", ""),
        # Организация — details
        "{organization}": organizations.get(org_id, {}).get("details", ""),
        "{customer}": customers.get(cust_id, {}).get("name", ""),
        "{season}": seasons.get(season_id, {}).get("name", ""),

        # Дистанция + Диспетчер из настроек/пейлоада
        "{distance}": fmt_num(distance_val),
        "{dispatcher}": dispatcher_val or "",

        # Даты
        "{dataPOPL}": dataPOPL,
        "{data_to}": data_to,
        "{dataSDPL}": dataSDPL,
        "{loading_time}": loading_time,
        "{unloading_time}": unloading_time,

        "{numberTN}": numberTN,
        "{tonn}": fmt_num(tonn),
        "{fuel_consumption}": fmt_num(fuel_consumption),

        "{dispatch_info}": dispatch_info,
        "{comment}": comment,

        "{driver_or_mechanic}": render_driver_or_mechanic(driver2),
    }

    return context


def fill_template_and_save(context, out_name, output_dir: Path | None = None):
    """Заполняет шаблон и сохраняет в Путевые листы\out_name. Возвращает путь."""
    if not PL_TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Шаблон не найден: {PL_TEMPLATE_PATH}")

    base_dir = output_dir if output_dir else get_default_output_dir()
    base_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(PL_TEMPLATE_PATH)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{" in cell.value and "}" in cell.value:
                    new_val = cell.value
                    for k, v in context.items():
                        if k in new_val:
                            new_val = new_val.replace(k, v if v is not None else "")
                    cell.value = new_val

    out_path = base_dir / out_name
    wb.save(out_path)
    return out_path
